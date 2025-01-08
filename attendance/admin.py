from datetime import time

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from django.contrib import admin
from django.db import connection
from django.http import HttpResponse
from django.urls import reverse
from django.utils.html import format_html
from django.utils.safestring import mark_safe

from attendance.models import Attendance

# Register your models here.
admin.site.site_header = 'Ishchilarning kelib ketish sistemase'
admin.site.site_title = 'Ishchilarning kelib ketish sistemase'
admin.site.index_title = 'Ishchilarning kelib ketish sistemase'


@admin.register(Attendance)
class QuestionAdmin(admin.ModelAdmin):
    # list_display = ['device_id', 'name', 'date', 'time', 'color_status']
    list_display = ['device_id', 'name', 'date', 'min_in_time', 'max_out_time', 'working_time']
    search_fields = ['name', 'device_id']
    # list_filter = ['status_color', 'is_in']
    ordering = ['-date', '-time']
    list_per_page = 15
    list_max_show_all = 100
    date_hierarchy = 'date'
    actions = ['download_excel']
    readonly_fields = ['device_id']

    # list_editable = ['status_color']

    # if user is not superuser remove clickable field
    def get_readonly_fields(self, request, obj=None):
        if not request.user.is_superuser:
            return [f.name for f in self.model._meta.fields]
        return []

    def download_excel(self, request, queryset):
        """
        Выгружает данные в формат XLSX:
        - №
        - FIO
        - Sana (дата)
        - Keldi (приход)
        - Ketdi (уход)
        - Ish vaqti (соат)
        с логикой цветов:
            - Keldi красный, если > 9:00, иначе зелёный
            - Ketdi красный, если < 18:00, иначе зелёный
            - Ish vaqti красный, если < 8, иначе зелёный
        """
        # 1) Создаём рабочую книгу
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance"

        # 2) Шапка
        headers = ["№", "FIO", "Sana", "Keldi", "Ketdi", "Ish vaqti (soat)"]
        ws.append(headers)

        # Настроим для шапки жирный шрифт, например
        bold_font = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="D9D9D9")

        for col_index, cell in enumerate(ws[1], start=1):
            cell.font = bold_font
            cell.fill = header_fill
            # Автоширина (немного «хак»: потом можно подобрать свои размеры)
            ws.column_dimensions[get_column_letter(col_index)].width = 18

        # 3) Проходимся по queryset и заполняем строки
        # Нам нужны те же вычисления, что и в админке: min_in_time, max_out_time, working_time.
        # Можно вынести эти методы в модель, чтобы не дублировать код, но для примера сделаем здесь.

        # Вспомогательная функция для перевода time -> seconds
        def time_to_seconds(t: time) -> int:
            return t.hour * 3600 + t.minute * 60 + t.second

        row_num = 2  # начинаем со второй строки (первая - заголовки)

        for index, item in enumerate(queryset, start=1):
            # ---- Получаем поля ----
            fio = item.name
            sana = item.date.strftime("%d-%m-%Y") if item.date else ""

            # --- min_in_time ---
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT MIN(b.time)
                    FROM public.attendance AS b
                    WHERE is_in = True AND b.device_id = %s AND b.date = %s
                """, [item.device_id, item.date])
                row_in = cursor.fetchone()
            min_time = row_in[0] if row_in and row_in[0] else None
            # Обрезаем время прихода к 9:00, если < 9:00
            # if min_time and min_time < time(9, 0):
            #     min_time = time(9, 0)

            # --- max_out_time ---
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT MAX(b.time)
                    FROM public.attendance AS b
                    WHERE is_in = False AND b.device_id = %s AND b.date = %s
                """, [item.device_id, item.date])
                row_out = cursor.fetchone()
            max_time = row_out[0] if row_out and row_out[0] else None
            # Обрезаем время ухода к 18:00, если > 18:00
            # if max_time and max_time > time(18, 0):
            #     max_time = time(18, 0)

            # Превращаем в строки для Excel
            if min_time:
                keldi_str = min_time.strftime('%H:%M')
            else:
                keldi_str = '-'
            if max_time:
                ketdi_str = max_time.strftime('%H:%M')
            else:
                ketdi_str = '-'

            # ---- working_time ----
            if min_time and max_time:
                diff_sec = time_to_seconds(max_time) - time_to_seconds(min_time)
                if diff_sec < 0:
                    diff_sec = 0
                hours = diff_sec // 3600
                minutes = (diff_sec % 3600) // 60
                ish_time_str = f"{hours}:{minutes:02d}"
            else:
                ish_time_str = "-"

            # 4) Записываем всё в XLS
            row_values = [
                index,  # №
                fio,  # FIO
                sana,  # Sana
                keldi_str,  # Keldi
                ketdi_str,  # Ketdi
                ish_time_str  # Ish vaqti (soat)
            ]
            ws.append(row_values)

            # 5) Логика по цветам
            #    Для ячеек Keldi (колонка 4), Ketdi (5), Ish vaqti (6)
            keldi_cell = ws.cell(row=row_num, column=4)
            ketdi_cell = ws.cell(row=row_num, column=5)
            ish_cell = ws.cell(row=row_num, column=6)

            # Keldi: красный, если > 9:00
            if min_time == "-":
                pass
            elif min_time and min_time > time(9, 0):
                keldi_cell.fill = PatternFill("solid", fgColor="FFC7CE")  # розовато-красный
            else:
                # Можно зелёный
                keldi_cell.fill = PatternFill("solid", fgColor="C6EFCE")

            # Ketdi: красный, если < 18:00 (и не '-')
            if max_time == "-":
                pass
            elif max_time and max_time < time(18, 0):
                ketdi_cell.fill = PatternFill("solid", fgColor="FFC7CE")
            else:
                ketdi_cell.fill = PatternFill("solid", fgColor="C6EFCE")

            # Ish vaqti: если >= 8 часов — зелёный, иначе красный
            if ish_time_str != "-":
                # hours:minutes
                parts = ish_time_str.split(":")
                if len(parts) == 2:
                    h = int(parts[0])
                    # m = int(parts[1]) -- минута при желании
                    if h >= 8:
                        ish_cell.fill = PatternFill("solid", fgColor="C6EFCE")  # зелёный
                    else:
                        ish_cell.fill = PatternFill("solid", fgColor="FFC7CE")  # красный
            # Иначе "-" не красим или как-то по-другому

            row_num += 1

        # 6) Готовим HttpResponse с XLSX
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="attendance.xlsx"'

        wb.save(response)
        return response

    # download_excel.short_description = "Выгрузить в Excel (XLSX)"
    download_excel.short_description = 'Excelga yuklash (XLSX)'

    def color_status(self, obj):
        if obj.status_color == 'red':
            # show red circle if time > 09:01:00
            return mark_safe('<span style="color: red;">&#11044;</span>')
        elif obj.status_color == 'yellow':
            return mark_safe('<span style="color: yellow;">&#11044;</span>')
        else:
            return mark_safe('<span style="color: green;">&#11044;</span>')

    color_status.short_description = 'Status'
    readonly_fields = ['color_status']

    def min_in_time(self, obj):
        """
        Возвращает минимальное время входа с цветовой индикацией:
        > 09:00 -> красным, иначе зелёным
        """
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT MIN(b.time)
                FROM public.attendance AS b
                WHERE is_in = True AND b.device_id = %s AND b.date = %s
            """, [obj.device_id, obj.date])
            row = cursor.fetchone()

        if not row or not row[0]:
            # Если нет данных, показываем прочерк
            return format_html('<span>-</span>')

        min_in_url = reverse('admin:attendance_attendance_change', args=[obj.pk])

        # row[0] предполагается объектом datetime.time (или datetime.datetime)
        min_time = row[0]

        # Условие: если пришёл позже 9:00
        if min_time > time(9, 0):
            color = "red"
        else:
            color = "green"

        return format_html(
            '<a href="{}" style="color: {};">{}</a>',
            min_in_url,
            color,
            min_time.strftime('%H:%M')
        )

    min_in_time.short_description = 'Keldi'

    def max_out_time(self, obj):
        """
        Возвращает максимальное время выхода с цветовой индикацией:
        < 18:00 -> красным, иначе зелёным
        """
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT MAX(b.time)
                FROM public.attendance AS b
                WHERE is_in = False AND b.device_id = %s AND b.date = %s
            """, [obj.device_id, obj.date])
            row = cursor.fetchone()

        if not row or not row[0]:
            return format_html('<span>-</span>')

        max_out_url = reverse('admin:attendance_attendance_change', args=[obj.pk])
        max_time = row[0]

        # Условие: если ушёл раньше 18:00
        if max_time < time(18, 0):
            color = "red"
        else:
            color = "green"

        return format_html(
            '<a href="{}" style="color: {};">{}</a>',
            max_out_url,
            color,
            max_time.strftime('%H:%M')
        )

    max_out_time.short_description = 'Ketdi'

    def changelist_view(self, request, extra_context=None):
        if not request.GET.get('is_in__exact'):
            q = request.GET.copy()
            q['is_in__exact'] = 'True'
            request.GET = q
            request.META['QUERY_STRING'] = request.GET.urlencode()
        return super().changelist_view(request, extra_context=extra_context)

    def working_time(self, obj):
        """
        Возвращает рабочее время в формате "ЧЧ:ММ".
        1) Обрезаем (09:00 .. 18:00).
        2) Считаем разницу прихода и ухода.
        3) Если > 6ч, вычитаем 1 час на обед (пример).
        4) Выводим HH:MM.
        """
        from datetime import time
        from django.db import connection
        from django.utils.html import format_html

        # 1. Получаем min_time (приход)
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT MIN(b.time)
                FROM public.attendance AS b
                WHERE is_in = True AND b.device_id = %s AND b.date = %s
            """, [obj.device_id, obj.date])
            row_in = cursor.fetchone()
        min_time = row_in[0] if row_in and row_in[0] else None

        # 2. Получаем max_time (уход)
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT MAX(b.time)
                FROM public.attendance AS b
                WHERE is_in = False AND b.device_id = %s AND b.date = %s
            """, [obj.device_id, obj.date])
            row_out = cursor.fetchone()
        max_time = row_out[0] if row_out and row_out[0] else None

        # Нет данных?
        if not min_time or not max_time:
            return format_html("<span>-</span>")

        # "Обрезаем" время
        if min_time < time(9, 0):
            min_time = time(9, 0)
        if max_time > time(18, 0):
            max_time = time(18, 0)

        # Переводим time -> секунды
        def time_to_seconds(t: time) -> int:
            return t.hour * 3600 + t.minute * 60 + t.second

        in_sec = time_to_seconds(min_time)
        out_sec = time_to_seconds(max_time)
        diff_sec = out_sec - in_sec

        if diff_sec <= 0:
            return format_html("<span>-</span>")

        # --- Если нужно учесть обед (пример) ---
        # Если работал больше 6 часов, вычитаем 1 час:
        if diff_sec >= 6 * 3600:
            diff_sec -= 3600  # вычитание часа

        # Часы и минуты
        hours = diff_sec // 3600
        minutes = (diff_sec % 3600) // 60

        # Пример форматирования: "7:59"
        diff_str = f"{hours}:{minutes:02d}"

        # Хотите подсветить по цвету? Например, зелёный, если >= 8 часов:
        # (учтите, что теперь мы уже вычли обед)
        color = "green" if (hours >= 8) else "red"

        return format_html(
            '<span style="color: {};">{}</span>',
            color,
            diff_str
        )

    working_time.short_description = 'Ish vaqti (soat)'
