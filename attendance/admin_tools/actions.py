from datetime import time
from io import BytesIO

import openpyxl
import pandas as pd
from django.db import connection
from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from attendance.models import Employee


def attendance_report(self, request, queryset):
    """
    Выгружает данные в формат XLSX:
    - №
    - FIO
    - Sana (дата)
    - Keldi (приход)
    - Ketdi (уход)
    - Ish vaqti (соат)
    с логикой цветов:
        - Keldi красный, если > 9:00, иначе зелёный
        - Ketdi красный, если < 18:00, иначе зелёный
        - Ish vaqti красный, если < 8, иначе зелёный
    """
    # 1) Создаём рабочую книгу
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # 2) Шапка
    headers = ["№", "FIO", "Sana", "Keldi", "Ketdi", "Ish vaqti (soat)", "Izoh"]
    ws.append(headers)

    # Настроим для шапки жирный шрифт, например
    bold_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9D9D9")

    for col_index, cell in enumerate(ws[1], start=1):
        cell.font = bold_font
        cell.fill = header_fill
        # Автоширина (немного «хак»: потом можно подобрать свои размеры)
        ws.column_dimensions[get_column_letter(col_index)].width = 18

    # 3) Проходимся по queryset и заполняем строки
    # Нам нужны те же вычисления, что и в админке: min_in_time, max_out_time, working_time.
    # Можно вынести эти методы в модель, чтобы не дублировать код, но для примера сделаем здесь.

    # Вспомогательная функция для перевода time -> seconds
    def time_to_seconds(t: time) -> int:
        return t.hour * 3600 + t.minute * 60 + t.second

    row_num = 2  # начинаем со второй строки (первая - заголовки)

    for index, item in enumerate(queryset, start=1):
        # ---- Получаем поля ----
        fio = item.name
        sana = item.date.strftime("%d-%m-%Y") if item.date else ""

        # --- min_in_time ---
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT MIN(b.time)
                FROM public.attendance AS b
                WHERE is_in = True AND b.device_id = %s AND b.date = %s
            """, [item.device_id, item.date])
            row_in = cursor.fetchone()
        min_time = row_in[0] if row_in and row_in[0] else None
        # Обрезаем время прихода к 9:00, если < 9:00
        # if min_time and min_time < time(9, 0):
        #     min_time = time(9, 0)

        # --- max_out_time ---
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT MAX(b.time)
                FROM public.attendance AS b
                WHERE is_in = False AND b.device_id = %s AND b.date = %s
            """, [item.device_id, item.date])
            row_out = cursor.fetchone()
        max_time = row_out[0] if row_out and row_out[0] else None
        # Обрезаем время ухода к 18:00, если > 18:00
        # if max_time and max_time > time(18, 0):
        #     max_time = time(18, 0)

        # Превращаем в строки для Excel
        if min_time:
            keldi_str = min_time.strftime('%H:%M')
        else:
            keldi_str = '-'
        if max_time:
            ketdi_str = max_time.strftime('%H:%M')
        else:
            ketdi_str = '-'

        # ---- working_time ----
        if min_time and max_time:
            diff_sec = time_to_seconds(max_time) - time_to_seconds(min_time)
            if diff_sec < 0:
                diff_sec = 0
            hours = diff_sec // 3600
            minutes = (diff_sec % 3600) // 60
            ish_time_str = f"{hours}:{minutes:02d}"
        else:
            ish_time_str = "-"

        # 4) Записываем всё в XLS
        row_values = [
            index,  # №
            fio,  # FIO
            sana,  # Sana
            keldi_str,  # Keldi
            ketdi_str,  # Ketdi
            ish_time_str,  # Ish vaqti (soat)
            item.description,  # Izoh
        ]
        ws.append(row_values)

        # 5) Логика по цветам
        #    Для ячеек Keldi (колонка 4), Ketdi (5), Ish vaqti (6)
        keldi_cell = ws.cell(row=row_num, column=4)
        ketdi_cell = ws.cell(row=row_num, column=5)
        ish_cell = ws.cell(row=row_num, column=6)

        # Keldi: красный, если > 9:00
        if min_time == "-":
            pass
        elif min_time and min_time > time(9, 5):
            keldi_cell.fill = PatternFill("solid", fgColor="FFC7CE")  # розовато-красный
        else:
            # Можно зелёный
            keldi_cell.fill = PatternFill("solid", fgColor="C6EFCE")

        # Ketdi: красный, если < 18:00 (и не '-')
        if max_time == "-":
            pass
        elif max_time and max_time < time(18, 0):
            ketdi_cell.fill = PatternFill("solid", fgColor="FFC7CE")
        else:
            ketdi_cell.fill = PatternFill("solid", fgColor="C6EFCE")

        # Ish vaqti: если >= 8 часов — зелёный, иначе красный
        if ish_time_str != "-":
            # hours:minutes
            parts = ish_time_str.split(":")
            if len(parts) == 2:
                h = int(parts[0])
                # m = int(parts[1]) -- минута при желании
                if h >= 8:
                    ish_cell.fill = PatternFill("solid", fgColor="C6EFCE")  # зелёный
                else:
                    ish_cell.fill = PatternFill("solid", fgColor="FFC7CE")  # красный
        # Иначе "-" не красим или как-то по-другому

        row_num += 1

    # 6) Готовим HttpResponse с XLSX
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # Получение даты из параметров
    selected_date = f"{request.GET.get('date__year', 'unknown')}-{request.GET.get('date__month', 'unknown')}-{request.GET.get('date__day', 'unknown')}"
    if selected_date == "unknown-unknown-unknown":
        selected_date = "attendance_report.xlsx"  # Если дата не указана, использовать дефолтное имя
    else:
        selected_date += "_attendance_report.xlsx"

    response['Content-Disposition'] = 'attachment; filename=' + selected_date

    wb.save(response)
    return response


attendance_report.short_description = 'Davomat haqida hisobotni excelga yuklash (XLSX)'


def absent_report_by_pinfl(self, request, queryset):
    """
    Генерация отчета по отсутствующим сотрудникам на основе PINFL.
    """
    # Load employee data from JSON file
    # try:
    #     with open('scripts/employees_data.json', 'r') as file:
    #         employees_data = json.load(file)
    # except FileNotFoundError:
    #     self.message_user(request, "Файл employees_data.json не найден.", level='error')
    #     return

    # Load employee data from Employees model
    employees_data = {
        employee.pinfl: {
            'last_name': employee.last_name,
            'first_name': employee.first_name,
            'middle_name': employee.middle_name,
            'dob': employee.dob,
            'position': employee.description,
        }
        for employee in Employee.objects.all()
    }

    # Extract all employees' PINFLs and details
    all_employees = {pinfl: data for pinfl, data in employees_data.items()}

    # Get PINFLs of employees who attended
    attended_pinfls = set(queryset.values_list('pinfl', flat=True))

    # Identify absent employees by PINFL
    absent_employees = [
        {
            'PINFL': pinfl,
            'Familiya': data['last_name'],
            'Ism': data['first_name'],
            # 'Middle Name': data.get('middle_name', ''),
            # 'Date of Birth': data.get('dob', ''),
            'Lavozimi': data.get('position', ''),
        }
        for pinfl, data in all_employees.items()
        if pinfl not in attended_pinfls
    ]

    # Convert absent employees to DataFrame for export
    report_df = pd.DataFrame(absent_employees)

    # Save to Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        report_df.to_excel(writer, index=False, sheet_name='Absent Report')

    # Prepare HTTP response
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    # Generate filename
    selected_date = f"{request.GET.get('date__year', 'unknown')}-{request.GET.get('date__month', 'unknown')}-{request.GET.get('date__day', 'unknown')}"
    filename = f"{selected_date}_absent_report_by_pinfl.xlsx" if selected_date != "unknown-unknown-unknown" else "absent_report_by_pinfl.xlsx"

    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


absent_report_by_pinfl.short_description = "Ro'yxatdan o'tmangan ishchilar haqida hisobot (PINFL)"


def attendance_report_by_pinfl(self, request, queryset):
    """
    Generates an Excel attendance report including PINFL and validation with color coding:
    - Keldi: Red if > 09:00, Green otherwise.
    - Ketdi: Red if < 18:00, Green otherwise.
    - Ish vaqti: Red if < 8 hours, Green otherwise.
    """
    # Load employee data from the Employee model
    employees_data = {
        employee.pinfl: {
            'last_name': employee.last_name,
            'first_name': employee.first_name,
        }
        for employee in Employee.objects.all()
    }

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # Define headers
    headers = [
        "No", "PINFL", "Familiya", "Ism", "Sana", "Keldi", "Ketdi", "Ish vaqti (soat)", "Izoh"
    ]
    ws.append(headers)

    # Apply header styles
    bold_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    for col_index, cell in enumerate(ws[1], start=1):
        cell.font = bold_font
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(col_index)].width = 18

    # Helper function to convert time to seconds
    def time_to_seconds(t: time) -> int:
        return t.hour * 3600 + t.minute * 60 + t.second

    # Iterate through queryset
    row_num = 2  # Start filling rows after the header
    for index, item in enumerate(queryset, start=1):
        # Get fields from the Attendance model
        pinfl = item.pinfl or "-"
        date = item.date.strftime("%d-%m-%Y") if item.date else "-"
        description = item.description or "-"

        # Check min_in_time
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT MIN(b.time)
                FROM public.attendance AS b
                WHERE is_in = True AND b.pinfl = %s AND b.date = %s
                """,
                [item.pinfl, item.date]
            )
            min_time_row = cursor.fetchone()
        min_time = min_time_row[0] if min_time_row and min_time_row[0] else None
        check_in = min_time.strftime('%H:%M') if min_time else "-"

        # Check max_out_time
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT MAX(b.time)
                FROM public.attendance AS b
                WHERE is_in = False AND b.pinfl = %s AND b.date = %s
                """,
                [item.pinfl, item.date]
            )
            max_time_row = cursor.fetchone()
        max_time = max_time_row[0] if max_time_row and max_time_row[0] else None
        check_out = max_time.strftime('%H:%M') if max_time else "-"

        # Calculate working hours
        if min_time and max_time:
            # Adjust times to working hours
            min_time = max(min_time, time(9, 0))
            max_time = min(max_time, time(18, 0))

            diff_sec = time_to_seconds(max_time) - time_to_seconds(min_time)
            if diff_sec <= 0:
                diff_sec = 0
            elif diff_sec >= 6 * 3600:
                diff_sec -= 3600

            hours = diff_sec // 3600
            minutes = (diff_sec % 3600) // 60
            working_hours = f"{hours}:{minutes:02d}"
        else:
            working_hours = "-"

        # Verify PINFL data
        last_name = employees_data.get(pinfl, {}).get('last_name', "-")
        first_name = employees_data.get(pinfl, {}).get('first_name', "-")

        # Append data to the worksheet
        row_values = [
            index, pinfl, last_name, first_name, date, check_in, check_out, working_hours, description
        ]
        ws.append(row_values)

        # Apply color coding
        keldi_cell = ws.cell(row=row_num, column=6)  # Keldi
        ketdi_cell = ws.cell(row=row_num, column=7)  # Ketdi
        ish_cell = ws.cell(row=row_num, column=8)  # Ish vaqti

        # Keldi: Red if > 09:00
        if min_time and min_time > time(9, 5):
            keldi_cell.fill = PatternFill("solid", fgColor="FFC7CE")  # Red
        elif min_time:
            keldi_cell.fill = PatternFill("solid", fgColor="C6EFCE")  # Green

        # Ketdi: Red if < 18:00
        if max_time and max_time < time(18, 0):
            ketdi_cell.fill = PatternFill("solid", fgColor="FFC7CE")  # Red
        elif max_time:
            ketdi_cell.fill = PatternFill("solid", fgColor="C6EFCE")  # Green

        # Ish vaqti: Red if < 8 hours
        if working_hours != "-":
            h, m = map(int, working_hours.split(":"))
            if h < 8:
                ish_cell.fill = PatternFill("solid", fgColor="FFC7CE")  # Red
            else:
                ish_cell.fill = PatternFill("solid", fgColor="C6EFCE")  # Green

        row_num += 1

    # Prepare Excel response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # Generate filename
    selected_date = f"{request.GET.get('date__year', 'unknown')}-{request.GET.get('date__month', 'unknown')}-{request.GET.get('date__day', 'unknown')}"
    filename = f"{selected_date}_attendance_report_with_pinfl.xlsx" if selected_date != "unknown-unknown-unknown" else "attendance_report_with_pinfl.xlsx"

    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


attendance_report_by_pinfl.short_description = "Davomat hisobotini PINFL bilan yuklash (XLSX)"
